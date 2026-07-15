from __future__ import annotations

import csv
import io
import os
import unicodedata
from collections import Counter
from typing import Any

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction

from apps.accounts.services.temporary_passwords import generate_temporary_password
from apps.accounts.models import User
from apps.audit.services.events import record_audit_event

IMPORT_MODE_CREATE_ONLY = "create_only"
IMPORT_MODE_UPDATE_EXISTING = "update_existing"
IMPORT_MODE_UPSERT = "upsert"
IMPORT_MODES = {IMPORT_MODE_CREATE_ONLY, IMPORT_MODE_UPDATE_EXISTING, IMPORT_MODE_UPSERT}


def _normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    return ascii_value.strip().lower().replace(" ", "_").replace("-", "_")


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_csv(uploaded_file) -> list[dict[str, str]]:
    raw = uploaded_file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(raw))
    return [
        {_normalize_key(key): _clean(value) for key, value in row.items() if key is not None}
        for row in reader
    ]


def _read_xlsx(uploaded_file) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Para importar Excel falta instalar openpyxl en el backend.") from exc

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_key(_clean(value)) for value in rows[0]]
    parsed_rows = []
    for values in rows[1:]:
        parsed_rows.append(
            {
                headers[index]: _clean(value)
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
        )
    return parsed_rows


def parse_user_import_file(uploaded_file) -> list[dict[str, str]]:
    extension = os.path.splitext(uploaded_file.name.lower())[1]
    if extension == ".csv":
        return _read_csv(uploaded_file)
    if extension == ".xlsx":
        return _read_xlsx(uploaded_file)
    raise ValueError("El archivo debe ser CSV o Excel .xlsx.")


def _value(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = _clean(row.get(_normalize_key(key), ""))
        if value:
            return value
    return ""


def _existing_user(*, email: str, legajo: str) -> tuple[User | None, str]:
    by_email = User.objects.filter(email__iexact=email).first() if email else None
    by_legajo = User.objects.filter(employee_code__iexact=legajo).first() if legajo else None
    if by_email and by_legajo and by_email.pk != by_legajo.pk:
        return None, "El email y el legajo pertenecen a usuarios distintos."
    return by_email or by_legajo, ""


def _username_for(email: str, legajo: str, requested_username: str = "") -> str:
    base = requested_username or email.split("@", 1)[0] or legajo
    base = _normalize_key(base).replace("_", ".")
    username = base or "usuario"
    suffix = 1
    while User.objects.filter(username__iexact=username).exists():
        username = f"{base}.{suffix}"
        suffix += 1
    return username


def _validate_row(row: dict[str, str], *, legajo_counts: Counter, email_counts: Counter, username_counts: Counter) -> tuple[dict[str, Any], list[str]]:
    legajo = _value(row, "legajo")
    first_name = _value(row, "nombre")
    last_name = _value(row, "apellido")
    email = _value(row, "email", "e_mail", "correo", "correo_electronico").lower()
    username = _value(row, "usuario", "username")
    phone = _value(row, "celular", "telefono", "phone")
    errors = []

    if legajo and legajo_counts[legajo.lower()] > 1:
        errors.append("Legajo repetido dentro del archivo.")

    if not email:
        errors.append("El email es obligatorio.")
    else:
        try:
            validate_email(email)
        except ValidationError:
            errors.append("Email invalido.")
        if email_counts[email] > 1:
            errors.append("Email repetido dentro del archivo.")

    if not first_name:
        errors.append("El nombre es obligatorio.")
    if not last_name:
        errors.append("El apellido es obligatorio.")

    existing, existing_error = _existing_user(email=email, legajo=legajo)
    if existing_error:
        errors.append(existing_error)
    if username and username_counts[username.lower()] > 1:
        errors.append("Usuario repetido dentro del archivo.")
    if username:
        username_owner = User.objects.filter(username__iexact=username).first()
        if username_owner and (not existing or username_owner.pk != existing.pk):
            errors.append("El usuario ya existe en otra cuenta.")

    return (
        {
            "legajo": legajo,
            "nombre": first_name,
            "apellido": last_name,
            "email": email,
            "usuario": username,
            "celular": phone,
            "existing": existing,
        },
        errors,
    )


def _counters(rows: list[dict[str, str]]) -> tuple[Counter, Counter, Counter]:
    legajos = Counter(_value(row, "legajo").lower() for row in rows if _value(row, "legajo"))
    emails = Counter(_value(row, "email", "e_mail", "correo", "correo_electronico").lower() for row in rows if _value(row, "email", "e_mail", "correo", "correo_electronico"))
    usernames = Counter(_value(row, "usuario", "username").lower() for row in rows if _value(row, "usuario", "username"))
    return legajos, emails, usernames


def analyze_user_import(uploaded_file, mode: str = IMPORT_MODE_UPSERT) -> dict[str, Any]:
    if mode not in IMPORT_MODES:
        raise ValueError("Modo de importacion invalido.")

    rows = parse_user_import_file(uploaded_file)
    legajo_counts, email_counts, username_counts = _counters(rows)
    items = []

    for index, row in enumerate(rows, start=2):
        data, errors = _validate_row(
            row,
            legajo_counts=legajo_counts,
            email_counts=email_counts,
            username_counts=username_counts,
        )
        existing = data.pop("existing")
        warnings = []
        action = "create"

        if existing:
            action = "update"
            if mode == IMPORT_MODE_CREATE_ONLY:
                action = "skip"
                warnings.append("Usuario existente: se omitira por el modo crear solo nuevos.")
        elif mode == IMPORT_MODE_UPDATE_EXISTING:
            action = "skip"
            warnings.append("Usuario nuevo: se omitira por el modo actualizar existentes.")

        items.append(
            {
                "row_number": index,
                **data,
                "existing_user_id": str(existing.pk) if existing else "",
                "status": "error" if errors else action,
                "errors": errors,
                "warnings": warnings,
            }
        )

    return {
        "mode": mode,
        "summary": {
            "total": len(items),
            "new_users": sum(1 for item in items if item["status"] == "create"),
            "existing_users": sum(1 for item in items if item["status"] == "update"),
            "errors": sum(1 for item in items if item["status"] == "error"),
            "skipped": sum(1 for item in items if item["status"] == "skip"),
            "duplicate_emails": sum(1 for count in email_counts.values() if count > 1),
            "duplicate_legajos": sum(1 for count in legajo_counts.values() if count > 1),
            "duplicate_usernames": sum(1 for count in username_counts.values() if count > 1),
        },
        "items": items,
    }


@transaction.atomic
def confirm_user_import(*, uploaded_file, mode: str, actor: User, request_id: str = "") -> dict[str, Any]:
    rows = parse_user_import_file(uploaded_file)
    legajo_counts, email_counts, username_counts = _counters(rows)
    created = 0
    updated = 0
    skipped = 0
    errors = []
    warnings = []
    report_items = []

    for index, row in enumerate(rows, start=2):
        data, item_errors = _validate_row(
            row,
            legajo_counts=legajo_counts,
            email_counts=email_counts,
            username_counts=username_counts,
        )
        existing = data["existing"]
        initial_password = None

        if item_errors:
            skipped += 1
            errors.extend([f"Fila {index}: {error}" for error in item_errors])
            status = "error"
        elif existing and mode == IMPORT_MODE_CREATE_ONLY:
            skipped += 1
            status = "skip"
            warnings.append(f"Fila {index}: usuario existente omitido.")
        elif not existing and mode == IMPORT_MODE_UPDATE_EXISTING:
            skipped += 1
            status = "skip"
            warnings.append(f"Fila {index}: usuario nuevo omitido.")
        else:
            before_data = {}
            if existing:
                user = existing
                before_data = {
                    "username": user.username,
                    "email": user.email,
                    "employee_code": user.employee_code,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "phone": user.phone,
                }
            else:
                user = User(username=_username_for(data["email"], data["legajo"], data["usuario"]), email=data["email"])
                user.access_level = User.AccessLevel.USUARIO_ACTIVO
                user.is_staff = False
                user.is_superuser = False
                user.is_active = True
                user.must_change_password = True
                user.password_changed_at = None
                initial_password = generate_temporary_password()
                user.set_password(initial_password)

            if data["usuario"] and user.username.lower() != data["usuario"].lower():
                user.username = data["usuario"]
            user.email = data["email"]
            if data["legajo"]:
                user.employee_code = data["legajo"]
            user.first_name = data["nombre"]
            user.last_name = data["apellido"]
            if data["celular"]:
                user.phone = data["celular"]
            user.save()

            record_audit_event(
                entity=user,
                action="user_bulk_import_update" if existing else "user_bulk_import_create",
                actor=actor,
                request_id=request_id,
                before_data=before_data,
                after_data={
                    "username": user.username,
                    "email": user.email,
                    "employee_code": user.employee_code,
                    "first_name": user.first_name,
                    "last_name": user.last_name,
                    "phone": user.phone,
                },
            )
            if existing:
                updated += 1
                status = "update"
            else:
                created += 1
                status = "create"

        report_items.append(
            {
                "row_number": index,
                "legajo": data["legajo"],
                "nombre": data["nombre"],
                "apellido": data["apellido"],
                "email": data["email"],
                "usuario": data["usuario"],
                "celular": data["celular"],
                "status": status,
                "errors": item_errors,
                "warnings": [],
                "initial_password": initial_password,
            }
        )

    record_audit_event(
        entity=actor,
        action="user_bulk_import",
        actor=actor,
        request_id=request_id,
        after_data={"created": created, "updated": updated, "skipped": skipped, "total": len(rows)},
    )
    return {
        "summary": {
            "total": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": len(errors),
            "warnings": len(warnings),
        },
        "items": report_items,
        "errors": errors,
        "warnings": warnings,
    }
